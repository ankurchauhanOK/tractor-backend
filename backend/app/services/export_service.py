import logging
from io import BytesIO
from typing import List

import pandas as pd
from fpdf import FPDF
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.database import Batch, Inspection
from app.services.storage import LocalStorage

logger = logging.getLogger(__name__)

storage = LocalStorage()


def _inspection_rows(inspections: List[Inspection]) -> List[dict]:
    rows = []
    for i in inspections:
        defects_text = "; ".join(d["text"] for d in (i.defects or []) if isinstance(d, dict))
        rows.append({
            "Page": i.page_number,
            "Tractor No": i.tractor_no,
            "Engine No": i.engine_no,
            "Chassis No": i.chassis_no,
            "Inspector": i.inspector,
            "Date": i.date.isoformat() if i.date else "",
            "Shift": i.shift,
            "Line No": i.line_no,
            "Defects": defects_text,
            "Status": i.status.value if i.status else "",
            "Needs Review": "Yes" if i.needs_review else "No",
            "Confidence": i.confidence_scores.get("tractor_no", "") if i.confidence_scores else "",
            "Verified By": i.verified_by,
            "Created At": i.created_at.isoformat() if i.created_at else "",
        })
    return rows


def generate_excel(batch_no: str, inspections: List[Inspection]) -> str:
    rows = _inspection_rows(inspections)
    df = pd.DataFrame(rows)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Inspections")
        ws = writer.sheets["Inspections"]

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

        for col_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

            col_width = max(len(str(col)) + 2, 12)
            for row_idx in range(2, len(df) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    col_width = max(col_width, min(len(str(val)) + 2, 60))
                if row_idx % 2 == 0:
                    ws.cell(row=row_idx, column=col_idx).fill = alt_fill

            ws.column_dimensions[get_column_letter(col_idx)].width = min(col_width, 60)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    buf.seek(0)
    filename = f"{batch_no}_inspection_report.xlsx"
    return storage.save_export(filename, buf.getvalue())


def generate_pdf(batch_no: str, batch: Batch, inspections: List[Inspection]) -> str:
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)

    # Title page
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 20)
    pdf.cell(0, 15, "Tractor Inspection Report", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, f"Batch: {batch_no}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Status: {batch.status.value if batch.status else 'N/A'}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Total Pages: {batch.total_pages}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Processed: {batch.processed_pages}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Verified: {batch.verified_pages}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Failed: {batch.failed_pages}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Avg Confidence: {batch.average_confidence}%" if batch.average_confidence else "Avg Confidence: N/A", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Factory: {batch.factory_name or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Plant: {batch.plant_name or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Line: {batch.line_name or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Operator: {batch.operator or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Created: {batch.created_at.strftime('%d/%m/%Y %H:%M') if batch.created_at else 'N/A'}", new_x="LMARGIN", new_y="NEXT")

    pdf.ln(10)

    # Summary table
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Inspection Details", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # Column headers
    pdf.set_font("Helvetica", "B", 8)
    col_widths = [10, 22, 22, 22, 20, 18, 14, 14, 50, 12]
    headers = ["Pg", "Tractor No", "Engine No", "Chassis No", "Inspector", "Date",
               "Shift", "Line", "Defects", "Status"]
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 7, h, border=1, align="C")
    pdf.ln()

    # Data rows
    pdf.set_font("Helvetica", "", 7)
    for inspection in inspections:
        defects = "; ".join(
            d["text"][:60] for d in (inspection.defects or [])
            if isinstance(d, dict)
        )
        row = [
            str(inspection.page_number or ""),
            inspection.tractor_no or "",
            inspection.engine_no or "",
            inspection.chassis_no or "",
            inspection.inspector or "",
            inspection.date.strftime("%d/%m/%Y") if inspection.date else "",
            inspection.shift or "",
            inspection.line_no or "",
            defects[:50],
            inspection.status.value if inspection.status else "",
        ]
        y_before = pdf.get_y()
        x_start = pdf.get_x()

        max_lines = 1
        for i, val in enumerate(row):
            lines = pdf.multi_cell(col_widths[i], 5, val, border=0, align="L", split_only=True)
            max_lines = max(max_lines, len(lines))

        row_h = max_lines * 5

        if y_before + row_h > pdf.h - 25:
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 8)
            for i2, h2 in enumerate(headers):
                pdf.cell(col_widths[i2], 7, h2, border=1, align="C")
            pdf.ln()
            pdf.set_font("Helvetica", "", 7)
            y_before = pdf.get_y()

        for i, val in enumerate(row):
            x = x_start + sum(col_widths[:i])
            pdf.set_xy(x, y_before)
            pdf.multi_cell(col_widths[i], 5, val, border=1, align="L")

        pdf.set_y(y_before + row_h)

    filename = f"{batch_no}_inspection_report.pdf"
    buf = pdf.output()
    return storage.save_export(filename, buf)
