import logging
from datetime import datetime
from io import BytesIO

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.models.database import Batch, Export, Inspection, get_db
from app.services.export_service import generate_excel, generate_pdf
from app.services.storage import storage

logger = logging.getLogger(__name__)

router = APIRouter()


class ExportRequest(BaseModel):
    format: str = "xlsx"


# ── Create export for a batch ─────────────────────────────────

@router.post("/batches/{batch_id}/exports", status_code=201)
def create_export(
    batch_id: int,
    body: ExportRequest,
    db: Session = Depends(get_db),
):
    if body.format not in ("xlsx", "pdf"):
        raise HTTPException(status_code=400, detail="Format must be 'xlsx' or 'pdf'")

    batch = db.query(Batch).filter(Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    inspections = (
        db.query(Inspection)
        .filter(Inspection.batch_id == batch_id)
        .order_by(Inspection.page_number)
        .all()
    )

    if not inspections:
        raise HTTPException(status_code=400, detail="Batch has no inspections to export")

    try:
        if body.format == "xlsx":
            file_path = generate_excel(batch.batch_no, inspections)
        else:
            file_path = generate_pdf(batch.batch_no, batch, inspections)
    except Exception as e:
        logger.error("Export generation failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

    export = Export(
        batch_id=batch.id,
        file_type=body.format,
        file_path=file_path,
        created_at=datetime.utcnow(),
    )
    db.add(export)
    db.commit()
    db.refresh(export)

    logger.info("Export %s created for batch %s", body.format, batch.batch_no)

    return {
        "id": export.id,
        "batch_id": export.batch_id,
        "batch_no": batch.batch_no,
        "file_type": export.file_type,
        "created_at": export.created_at.isoformat() if export.created_at else None,
    }


# ── List exports for a batch ──────────────────────────────────

@router.get("/batches/{batch_id}/exports")
def list_exports(
    batch_id: int,
    db: Session = Depends(get_db),
):
    batch = db.query(Batch).filter(Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    exports = (
        db.query(Export)
        .filter(Export.batch_id == batch_id)
        .order_by(Export.created_at.desc())
        .all()
    )

    return {
        "batch_id": batch_id,
        "batch_no": batch.batch_no,
        "exports": [e.to_dict() for e in exports],
    }


# ── Download export file ──────────────────────────────────────

@router.get("/exports/{export_id}/download")
def download_export(
    export_id: int,
    db: Session = Depends(get_db),
):
    export = db.query(Export).filter(Export.id == export_id).first()
    if not export:
        raise HTTPException(status_code=404, detail="Export not found")

    media_types = {
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "pdf": "application/pdf",
    }
    filename = export.file_path.split("/")[-1] or f"export_{export_id}"
    media_type = media_types.get(export.file_type, "application/octet-stream")

    data = storage.read_file_by_key(export.file_path)
    if data is None:
        raise HTTPException(status_code=404, detail="Export file not found in storage")

    return StreamingResponse(
        iter([data]),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Legacy: export all inspections (backward compat) ──────────

@router.get("/export")
def export_all_excel(db: Session = Depends(get_db)):
    inspections = db.query(Inspection).order_by(Inspection.created_at.asc(), Inspection.id.asc()).all()

    rows = []
    for i in inspections:
        defects_text = "; ".join([d["text"] for d in (i.defects or []) if isinstance(d, dict)])
        rows.append({
            "ID": i.id,
            "Tractor No": i.tractor_no,
            "Date": i.date,
            "Shift": i.shift,
            "Line No": i.line_no,
            "Defects": defects_text,
            "Verified By": i.verified_by,
            "Final Verified By": i.final_verified_by,
            "Status": i.status.value if i.status else "",
            "Created At": i.created_at.isoformat() if i.created_at else "",
        })

    df = pd.DataFrame(rows)
    filename = f"inspection_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Inspections")
        ws = writer.sheets["Inspections"]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        for col_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

            max_len = len(str(col))
            for row_idx in range(2, len(df) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    cell_len = len(str(val))
                    if cell_len > max_len:
                        max_len = cell_len
                if row_idx % 2 == 0:
                    ws.cell(row=row_idx, column=col_idx).fill = alt_fill

            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    buf.seek(0)
    object_key = storage.save_export(filename, buf.getvalue())

    data = storage.read_file_by_key(object_key)
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        iter([data]),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
